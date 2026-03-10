<<<<<<< HEAD
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import sqlite3
from datetime import datetime, date, timedelta
from werkzeug.security import check_password_hash, generate_password_hash
from openpyxl import Workbook
from io import BytesIO
import re

app = Flask(__name__)
app.secret_key = "your_secret_key_here"
DB_NAME = "company_data.db"


def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def login_required():
    return "admin_id" in session


def user_login_required():
    return "user_id" in session


def is_admin():
    return "admin_id" in session


def can_user_access_table(user_id, table_id):
    conn = get_db_connection()

    allowed = conn.execute("""
        SELECT 1
        FROM user_table_permissions
        WHERE user_id = ? AND table_id = ?
    """, (user_id, table_id)).fetchone()

    own_table = conn.execute("""
        SELECT 1
        FROM company_tables
        WHERE id = ? AND created_by_user_id = ?
    """, (table_id, user_id)).fetchone()

    conn.close()
    return allowed is not None or own_table is not None


def get_expiry_status(expiry_date_str):
    try:
        expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
        today = date.today()
        days_left = (expiry_date - today).days

        if days_left < 0:
            return "Expired", "danger", days_left
        elif days_left <= 7:
            return "Expiring Soon", "warning", days_left
        elif days_left <= 30:
            return "Near Expiry", "info", days_left
        else:
            return "Active", "success", days_left
    except Exception:
        return None, None, None


def get_alerts():
    conn = get_db_connection()

    query = """
    SELECT 
        c.company_name,
        ct.table_name,
        tr.id AS row_id,
        tc.column_name,
        rv.value AS expiry_value
    FROM row_values rv
    JOIN table_rows tr ON rv.row_id = tr.id
    JOIN table_columns tc ON rv.column_id = tc.id
    JOIN company_tables ct ON tr.table_id = ct.id
    JOIN companies c ON ct.company_id = c.id
    WHERE tc.is_expiry = 1
    """
    rows = conn.execute(query).fetchall()
    conn.close()

    alerts = []
    for row in rows:
        status, badge, days_left = get_expiry_status(row["expiry_value"])
        if status in ["Expired", "Expiring Soon", "Near Expiry"]:
            alerts.append({
                "company_name": row["company_name"],
                "table_name": row["table_name"],
                "column_name": row["column_name"],
                "expiry_value": row["expiry_value"],
                "status": status,
                "badge": badge,
                "days_left": days_left
            })

    alerts.sort(key=lambda x: x["days_left"] if x["days_left"] is not None else 99999)
    return alerts


def add_audit_log(action_type, description, company_id=None, company_name=None, table_id=None, table_name=None):
    conn = get_db_connection()

    dubai_time = (datetime.utcnow() + timedelta(hours=4)).strftime("%Y-%m-%d %H:%M:%S")

    if session.get("admin_id"):
        actor_role = "Admin"
        actor_username = session.get("username", "Unknown")
    elif session.get("user_id"):
        actor_role = "User"
        actor_username = session.get("user_username", "Unknown")
    else:
        actor_role = "Unknown"
        actor_username = "Unknown"

    conn.execute("""
        INSERT INTO audit_logs (
            actor_role,
            actor_username,
            company_id,
            company_name,
            table_id,
            table_name,
            action_type,
            description,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        actor_role,
        actor_username,
        company_id,
        company_name,
        table_id,
        table_name,
        action_type,
        description,
        dubai_time
    ))

    conn.commit()
    conn.close()


@app.route("/")
def home():
    return render_template("home.html")


# -------------------------
# ADMIN AUTH
# -------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        conn = get_db_connection()
        admin = conn.execute(
            "SELECT * FROM admins WHERE username = ?",
            (username,)
        ).fetchone()
        conn.close()

        if admin and check_password_hash(admin["password_hash"], password):
            session.clear()
            session["admin_id"] = admin["id"]
            session["username"] = admin["username"]
            flash("Login successful.", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password.", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for("login"))


# -------------------------
# USER AUTH
# -------------------------
@app.route("/user-login", methods=["GET", "POST"])
def user_login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        conn = get_db_connection()
        user = conn.execute(
            "SELECT * FROM users WHERE username = ?",
            (username,)
        ).fetchone()
        conn.close()

        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            session["user_username"] = user["username"]
            flash("User login successful.", "success")
            return redirect(url_for("user_dashboard"))
        else:
            flash("Invalid username or password.", "danger")

    return render_template("user_login.html")


@app.route("/user-logout")
def user_logout():
    session.clear()
    flash("User logged out successfully.", "info")
    return redirect(url_for("user_login"))


# -------------------------
# DASHBOARDS
# -------------------------
@app.route("/dashboard")
def dashboard():
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    total_companies = conn.execute("SELECT COUNT(*) AS count FROM companies").fetchone()["count"]
    total_tables = conn.execute("SELECT COUNT(*) AS count FROM company_tables").fetchone()["count"]
    conn.close()

    alerts = get_alerts()

    expired_count = sum(1 for a in alerts if a["status"] == "Expired")
    soon_count = sum(1 for a in alerts if a["status"] == "Expiring Soon")
    near_count = sum(1 for a in alerts if a["status"] == "Near Expiry")

    return render_template(
        "dashboard.html",
        total_companies=total_companies,
        total_tables=total_tables,
        alerts=alerts,
        expired_count=expired_count,
        soon_count=soon_count,
        near_count=near_count
    )


@app.route("/user-dashboard")
def user_dashboard():
    if not user_login_required():
        return redirect(url_for("user_login"))

    conn = get_db_connection()
    companies_data = conn.execute("SELECT * FROM companies ORDER BY id DESC").fetchall()
    conn.close()

    return render_template("companies.html", companies=companies_data)


# -------------------------
# COMPANIES
# -------------------------
@app.route("/companies", methods=["GET", "POST"])
def companies():
    if not login_required() and not user_login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        notes = request.form.get("notes", "").strip()

        if company_name:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO companies (company_name, notes) VALUES (?, ?)",
                (company_name, notes)
            )
            conn.commit()

            new_company_id = cursor.lastrowid

            add_audit_log(
                action_type="Add Company",
                description=f"Added company '{company_name}'",
                company_id=new_company_id,
                company_name=company_name
            )

            flash("Company added successfully.", "success")
        else:
            flash("Company name is required.", "danger")

    companies_data = conn.execute("SELECT * FROM companies ORDER BY id DESC").fetchall()
    conn.close()

    return render_template("companies.html", companies=companies_data)


@app.route("/company/<int:company_id>")
def company_detail(company_id):
    if not login_required() and not user_login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    if user_login_required() and not is_admin():
        user_id = session["user_id"]

        tables = conn.execute("""
            SELECT *
            FROM company_tables
            WHERE company_id = ?
            AND (
                id IN (
                    SELECT table_id
                    FROM user_table_permissions
                    WHERE user_id = ?
                )
                OR created_by_user_id = ?
            )
            ORDER BY id DESC
        """, (company_id, user_id, user_id)).fetchall()
    else:
        tables = conn.execute("""
            SELECT *
            FROM company_tables
            WHERE company_id = ?
            ORDER BY id DESC
        """, (company_id,)).fetchall()

    details = conn.execute("""
        SELECT * FROM company_details
        WHERE company_id = ?
        ORDER BY id DESC
    """, (company_id,)).fetchall()

    conn.close()

    return render_template(
        "company_detail.html",
        company=company,
        tables=tables,
        details=details
    )


@app.route("/company/edit/<int:company_id>", methods=["GET", "POST"])
def edit_company(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        notes = request.form.get("notes", "").strip()

        if not company_name:
            flash("Company name is required.", "danger")
            return render_template("edit_company.html", company=company)

        old_name = company["company_name"]

        conn.execute(
            "UPDATE companies SET company_name = ?, notes = ? WHERE id = ?",
            (company_name, notes, company_id)
        )

        conn.commit()
        conn.close()

        add_audit_log(
            action_type="Edit Company",
            description=f"Updated company '{old_name}' to '{company_name}'",
            company_id=company_id,
            company_name=company_name
        )

        flash("Company updated successfully.", "success")
        return redirect(url_for("companies"))

    conn.close()
    return render_template("edit_company.html", company=company)


@app.route("/company/delete/<int:company_id>")
def delete_company(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    company_name = company["company_name"] if company else "Unknown"

    conn.execute("DELETE FROM company_details WHERE company_id = ?", (company_id,))

    conn.execute("""
        DELETE FROM row_values
        WHERE row_id IN (
            SELECT id FROM table_rows
            WHERE table_id IN (
                SELECT id FROM company_tables WHERE company_id = ?
            )
        )
    """, (company_id,))

    conn.execute("""
        DELETE FROM table_rows
        WHERE table_id IN (
            SELECT id FROM company_tables WHERE company_id = ?
        )
    """, (company_id,))

    conn.execute("""
        DELETE FROM table_columns
        WHERE table_id IN (
            SELECT id FROM company_tables WHERE company_id = ?
        )
    """, (company_id,))

    conn.execute("DELETE FROM company_tables WHERE company_id = ?", (company_id,))
    conn.execute("DELETE FROM companies WHERE id = ?", (company_id,))

    conn.commit()
    conn.close()

    add_audit_log(
        action_type="Delete Company",
        description=f"Deleted company '{company_name}' and all related data",
        company_id=company_id,
        company_name=company_name
    )

    flash("Company deleted successfully.", "warning")
    return redirect(url_for("companies"))


# -------------------------
# COMPANY DETAILS
# -------------------------
@app.route("/company/<int:company_id>/details", methods=["GET", "POST"])
def company_details(company_id):
    if not login_required() and not user_login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    if request.method == "POST":
        detail_label = request.form.get("detail_label", "").strip()
        detail_description = request.form.get("detail_description", "").strip()

        if detail_label:
            conn.execute("""
                INSERT INTO company_details (company_id, detail_label, detail_value)
                VALUES (?, ?, ?)
            """, (company_id, detail_label, detail_description))

            conn.commit()

            add_audit_log(
                action_type="Add Company Detail",
                description=f"Added detail '{detail_label}' to company '{company['company_name']}'",
                company_id=company_id,
                company_name=company["company_name"]
            )

            flash("Company detail added successfully.", "success")
        else:
            flash("Detail label is required.", "danger")

    details = conn.execute("""
        SELECT * FROM company_details
        WHERE company_id = ?
        ORDER BY id DESC
    """, (company_id,)).fetchall()

    conn.close()

    return render_template(
        "company_details.html",
        company=company,
        details=details
    )


@app.route("/company/detail/edit/<int:detail_id>/<int:company_id>", methods=["GET", "POST"])
def edit_company_detail(detail_id, company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    detail = conn.execute(
        "SELECT * FROM company_details WHERE id = ?",
        (detail_id,)
    ).fetchone()

    if not detail:
        conn.close()
        flash("Detail not found.", "danger")
        return redirect(url_for("company_detail", company_id=company_id))

    if request.method == "POST":
        label = request.form.get("detail_label", "").strip()
        description = request.form.get("detail_description", "").strip()

        conn.execute(
            "UPDATE company_details SET detail_label = ?, detail_value = ? WHERE id = ?",
            (label, description, detail_id)
        )

        company = conn.execute(
            "SELECT * FROM companies WHERE id = ?",
            (company_id,)
        ).fetchone()

        conn.commit()
        conn.close()

        add_audit_log(
            action_type="Edit Company Detail",
            description=f"Updated detail '{label}' in company '{company['company_name']}'",
            company_id=company_id,
            company_name=company["company_name"]
        )

        flash("Company detail updated successfully.", "success")
        return redirect(url_for("company_detail", company_id=company_id))

    conn.close()
    return render_template(
        "edit_company_detail.html",
        detail=detail,
        company_id=company_id
    )


@app.route("/company/detail/delete/<int:detail_id>/<int:company_id>")
def delete_company_detail(detail_id, company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    detail = conn.execute(
        "SELECT * FROM company_details WHERE id = ?",
        (detail_id,)
    ).fetchone()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not detail:
        conn.close()
        flash("Company detail not found.", "danger")
        return redirect(url_for("company_detail", company_id=company_id))

    conn.execute("DELETE FROM company_details WHERE id = ?", (detail_id,))
    conn.commit()
    conn.close()

    add_audit_log(
        action_type="Delete Company Detail",
        description=f"Deleted detail '{detail['detail_label']}' from company '{company['company_name']}'",
        company_id=company_id,
        company_name=company["company_name"]
    )

    flash("Company detail deleted successfully.", "warning")
    return redirect(url_for("company_detail", company_id=company_id))


# -------------------------
# TABLES
# -------------------------
@app.route("/company/<int:company_id>/create_table", methods=["GET", "POST"])
def create_table(company_id):
    if not login_required() and not user_login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    users = conn.execute("""
        SELECT * FROM users
        ORDER BY username ASC
    """).fetchall()

    if request.method == "POST":
        table_name = request.form.get("table_name", "").strip()
        column_names = request.form.getlist("column_name[]")
        column_types = request.form.getlist("column_type[]")
        expiry_columns = request.form.getlist("is_expiry[]")
        allowed_users = request.form.getlist("allowed_users")

        if not table_name:
            flash("Table name is required.", "danger")
            conn.close()
            return render_template("create_table.html", company=company, users=users)

        cursor = conn.cursor()
        created_by_user_id = session.get("user_id") if "user_id" in session else None

        cursor.execute(
            "INSERT INTO company_tables (company_id, table_name, created_by_user_id) VALUES (?, ?, ?)",
            (company_id, table_name, created_by_user_id)
        )
        table_id = cursor.lastrowid

        display_order = 1
        for i in range(len(column_names)):
            name = column_names[i].strip()
            col_type = column_types[i].strip() if i < len(column_types) else "text"
            is_expiry = 1 if str(i) in expiry_columns else 0

            if name:
                cursor.execute("""
                    INSERT INTO table_columns (table_id, column_name, column_type, is_expiry, display_order)
                    VALUES (?, ?, ?, ?, ?)
                """, (table_id, name, col_type, is_expiry, display_order))
                display_order += 1

        if login_required():
            for user_id in allowed_users:
                cursor.execute("""
                    INSERT INTO user_table_permissions (user_id, table_id)
                    VALUES (?, ?)
                """, (user_id, table_id))

        conn.commit()
        conn.close()

        add_audit_log(
            action_type="Create Table",
            description=f"Created table '{table_name}' in company '{company['company_name']}'",
            company_id=company_id,
            company_name=company["company_name"],
            table_id=table_id,
            table_name=table_name
        )

        flash("Dynamic table created successfully.", "success")
        return redirect(url_for("company_detail", company_id=company_id))

    conn.close()
    return render_template(
        "create_table.html",
        company=company,
        users=users
    )


@app.route("/table/<int:table_id>")
def view_table(table_id):
    if not login_required() and not user_login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    if "user_id" in session and "admin_id" not in session:
        user_id = session["user_id"]

        allowed = conn.execute("""
            SELECT 1
            FROM user_table_permissions
            WHERE user_id = ? AND table_id = ?
        """, (user_id, table_id)).fetchone()

        owner = conn.execute("""
            SELECT 1
            FROM company_tables
            WHERE id = ? AND created_by_user_id = ?
        """, (table_id, user_id)).fetchone()

        if not allowed and not owner:
            conn.close()
            flash("You do not have permission to view this table.", "danger")
            return redirect(url_for("user_dashboard"))

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (table["company_id"],)
    ).fetchone()

    columns = conn.execute("""
        SELECT * FROM table_columns
        WHERE table_id = ?
        ORDER BY display_order ASC
    """, (table_id,)).fetchall()

    rows = conn.execute("""
        SELECT * FROM table_rows
        WHERE table_id = ?
        ORDER BY id DESC
    """, (table_id,)).fetchall()

    table_data = []
    for row in rows:
        values = conn.execute("""
            SELECT rv.value, tc.column_name, tc.is_expiry
            FROM row_values rv
            JOIN table_columns tc ON rv.column_id = tc.id
            WHERE rv.row_id = ?
            ORDER BY tc.display_order ASC
        """, (row["id"],)).fetchall()

        row_dict = {}
        row_status = None
        row_badge = None

        for value in values:
            row_dict[value["column_name"]] = value["value"]
            if value["is_expiry"] == 1:
                status, badge, _ = get_expiry_status(value["value"])
                row_status = status
                row_badge = badge

        table_data.append({
            "id": row["id"],
            "row_values": row_dict,
            "status": row_status,
            "badge": row_badge
        })

    conn.close()

    return render_template(
        "view_table.html",
        table=table,
        company=company,
        columns=columns,
        table_data=table_data
    )


@app.route("/table/<int:table_id>/add_row", methods=["GET", "POST"])
def add_row(table_id):
    if not login_required() and not user_login_required():
        return redirect(url_for("login"))

    if user_login_required() and not is_admin():
        if not can_user_access_table(session["user_id"], table_id):
            flash("You do not have permission for this table.", "danger")
            return redirect(url_for("user_dashboard"))

    conn = get_db_connection()

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    columns = conn.execute("""
        SELECT * FROM table_columns
        WHERE table_id = ?
        ORDER BY display_order ASC
    """, (table_id,)).fetchall()

    if request.method == "POST":
        cursor = conn.cursor()
        cursor.execute("INSERT INTO table_rows (table_id) VALUES (?)", (table_id,))
        row_id = cursor.lastrowid

        for column in columns:
            value = request.form.get(f"column_{column['id']}", "").strip()
            cursor.execute("""
                INSERT INTO row_values (row_id, column_id, value)
                VALUES (?, ?, ?)
            """, (row_id, column["id"], value))

        conn.commit()

        company = conn.execute(
            "SELECT * FROM companies WHERE id = ?",
            (table["company_id"],)
        ).fetchone()

        conn.close()

        add_audit_log(
            action_type="Add Row",
            description=f"Added a row to table '{table['table_name']}' in company '{company['company_name']}'",
            company_id=company["id"],
            company_name=company["company_name"],
            table_id=table_id,
            table_name=table["table_name"]
        )

        flash("Row added successfully.", "success")
        return redirect(url_for("view_table", table_id=table_id))

    conn.close()
    return render_template("add_row.html", table=table, columns=columns)


@app.route("/row/delete/<int:row_id>/<int:table_id>")
def delete_row(row_id, table_id):
    if not login_required() and not user_login_required():
        return redirect(url_for("login"))

    if user_login_required() and not is_admin():
        if not can_user_access_table(session["user_id"], table_id):
            flash("You do not have permission for this table.", "danger")
            return redirect(url_for("user_dashboard"))

    conn = get_db_connection()

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (table["company_id"],)
    ).fetchone()

    conn.execute("DELETE FROM row_values WHERE row_id = ?", (row_id,))
    conn.execute("DELETE FROM table_rows WHERE id = ?", (row_id,))
    conn.commit()
    conn.close()

    add_audit_log(
        action_type="Delete Row",
        description=f"Deleted a row from table '{table['table_name']}' in company '{company['company_name']}'",
        company_id=company["id"],
        company_name=company["company_name"],
        table_id=table_id,
        table_name=table["table_name"]
    )

    flash("Row deleted successfully.", "warning")
    return redirect(url_for("view_table", table_id=table_id))


@app.route("/table/delete/<int:table_id>")
def delete_table(table_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    company_id = table["company_id"]

    conn.execute("""
        DELETE FROM row_values
        WHERE row_id IN (
            SELECT id FROM table_rows WHERE table_id = ?
        )
    """, (table_id,))

    conn.execute("DELETE FROM table_rows WHERE table_id = ?", (table_id,))
    conn.execute("DELETE FROM table_columns WHERE table_id = ?", (table_id,))
    conn.execute("DELETE FROM company_tables WHERE id = ?", (table_id,))
    conn.commit()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    conn.close()

    add_audit_log(
        action_type="Delete Table",
        description=f"Deleted table '{table['table_name']}' from company '{company['company_name']}'",
        company_id=company_id,
        company_name=company["company_name"],
        table_id=table_id,
        table_name=table["table_name"]
    )

    flash("Table deleted successfully.", "warning")
    return redirect(url_for("company_detail", company_id=company_id))


# -------------------------
# EXPORTS
# -------------------------
@app.route("/company/<int:company_id>/export-full")
def export_full_company(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    workbook = Workbook()

    info_sheet = workbook.active
    info_sheet.title = "Company Info"
    info_sheet["A1"] = "Field"
    info_sheet["B1"] = "Value"

    row_num = 2
    for key in company.keys():
        info_sheet.cell(row=row_num, column=1, value=key)
        info_sheet.cell(row=row_num, column=2, value=company[key])
        row_num += 1

    info_sheet.column_dimensions["A"].width = 25
    info_sheet.column_dimensions["B"].width = 40

    details_sheet = workbook.create_sheet("Company Details")

    try:
        details = conn.execute("""
            SELECT * FROM company_details
            WHERE company_id = ?
            ORDER BY id DESC
        """, (company_id,)).fetchall()

        if details:
            headers = details[0].keys()
            for col_num, header in enumerate(headers, 1):
                details_sheet.cell(row=1, column=col_num, value=header)

            for row_index, detail in enumerate(details, start=2):
                for col_num, header in enumerate(headers, 1):
                    details_sheet.cell(row=row_index, column=col_num, value=detail[header])
        else:
            details_sheet["A1"] = "No company details found"

    except sqlite3.OperationalError:
        details_sheet["A1"] = "company_details table does not exist"

    for col in details_sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        details_sheet.column_dimensions[column_letter].width = max_length + 2

    tables = conn.execute("""
        SELECT * FROM company_tables
        WHERE company_id = ?
        ORDER BY id DESC
    """, (company_id,)).fetchall()

    if not tables:
        note_sheet = workbook.create_sheet("Tables")
        note_sheet["A1"] = "No company tables found"
    else:
        for table in tables:
            table_name = table["table_name"] if table["table_name"] else f"Table_{table['id']}"
            safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '', table_name)[:31]

            if not safe_sheet_name:
                safe_sheet_name = f"Table_{table['id']}"

            sheet = workbook.create_sheet(safe_sheet_name)

            columns = conn.execute("""
                SELECT * FROM table_columns
                WHERE table_id = ?
                ORDER BY display_order ASC
            """, (table["id"],)).fetchall()

            if not columns:
                sheet["A1"] = "No columns found for this table"
                continue

            headers = [column["column_name"] for column in columns]
            headers.append("Status")
            sheet.append(headers)

            rows = conn.execute("""
                SELECT * FROM table_rows
                WHERE table_id = ?
                ORDER BY id DESC
            """, (table["id"],)).fetchall()

            for row in rows:
                values = conn.execute("""
                    SELECT rv.value, tc.column_name, tc.is_expiry
                    FROM row_values rv
                    JOIN table_columns tc ON rv.column_id = tc.id
                    WHERE rv.row_id = ?
                    ORDER BY tc.display_order ASC
                """, (row["id"],)).fetchall()

                row_dict = {}
                row_status = ""

                for value in values:
                    row_dict[value["column_name"]] = value["value"]
                    if value["is_expiry"] == 1:
                        status, _, _ = get_expiry_status(value["value"])
                        row_status = status if status else ""

                excel_row = [row_dict.get(column["column_name"], "") for column in columns]
                excel_row.append(row_status)
                sheet.append(excel_row)

            for col in sheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2

    conn.close()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    company_name = company["company_name"].replace(" ", "_")
    filename = f"{company_name}_full_export.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/table/<int:table_id>/export")
def export_table_excel(table_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (table["company_id"],)
    ).fetchone()

    columns = conn.execute("""
        SELECT * FROM table_columns
        WHERE table_id = ?
        ORDER BY display_order ASC
    """, (table_id,)).fetchall()

    rows = conn.execute("""
        SELECT * FROM table_rows
        WHERE table_id = ?
        ORDER BY id DESC
    """, (table_id,)).fetchall()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Table Data"

    headers = [column["column_name"] for column in columns]
    headers.append("Status")
    worksheet.append(headers)

    for row in rows:
        values = conn.execute("""
            SELECT rv.value, tc.column_name, tc.is_expiry
            FROM row_values rv
            JOIN table_columns tc ON rv.column_id = tc.id
            WHERE rv.row_id = ?
            ORDER BY tc.display_order ASC
        """, (row["id"],)).fetchall()

        row_dict = {}
        row_status = ""

        for value in values:
            row_dict[value["column_name"]] = value["value"]
            if value["is_expiry"] == 1:
                status, _, _ = get_expiry_status(value["value"])
                row_status = status if status else ""

        excel_row = [row_dict.get(column["column_name"], "") for column in columns]
        excel_row.append(row_status)
        worksheet.append(excel_row)

    conn.close()

    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = max_length + 2

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    company_name = company["company_name"].replace(" ", "_") if company else "Company"
    table_name = table["table_name"].replace(" ", "_")
    filename = f"{company_name}_{table_name}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# -------------------------
# ADMIN USER MANAGEMENT
# -------------------------
@app.route("/admin/users/create", methods=["GET", "POST"])
def create_user():
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if not username or not password:
            flash("Username and password are required.", "danger")
            conn.close()
            return render_template("create_user.html")

        existing_user = conn.execute(
            "SELECT * FROM users WHERE username = ?",
            (username,)
        ).fetchone()

        if existing_user:
            flash("Username already exists.", "danger")
            conn.close()
            return render_template("create_user.html")

        password_hash = generate_password_hash(password)

        conn.execute("""
            INSERT INTO users (username, password_hash)
            VALUES (?, ?)
        """, (username, password_hash))
        conn.commit()
        conn.close()

        add_audit_log(
            action_type="Create User",
            description=f"Created user '{username}'"
        )

        flash("User created successfully.", "success")
        return redirect(url_for("manage_users"))

    conn.close()
    return render_template("create_user.html")


@app.route("/admin/users")
def manage_users():
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    users = conn.execute("""
        SELECT * FROM users
        ORDER BY id DESC
    """).fetchall()
    conn.close()

    return render_template("manage_users.html", users=users)


@app.route("/admin/users/<int:user_id>/permissions", methods=["GET", "POST"])
def user_permissions(user_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    user = conn.execute(
        "SELECT * FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not user:
        conn.close()
        flash("User not found.", "danger")
        return redirect(url_for("manage_users"))

    tables = conn.execute("""
        SELECT company_tables.*, companies.company_name
        FROM company_tables
        LEFT JOIN companies ON company_tables.company_id = companies.id
        ORDER BY company_tables.id DESC
    """).fetchall()

    if request.method == "POST":
        selected_tables = request.form.getlist("table_permissions")

        conn.execute(
            "DELETE FROM user_table_permissions WHERE user_id = ?",
            (user_id,)
        )

        for table_id in selected_tables:
            conn.execute("""
                INSERT INTO user_table_permissions (user_id, table_id)
                VALUES (?, ?)
            """, (user_id, table_id))

        conn.commit()

        add_audit_log(
            action_type="Update User Permissions",
            description=f"Updated table permissions for user '{user['username']}'"
        )

        flash("Permissions updated successfully.", "success")

    allowed_table_ids = [
        row["table_id"] for row in conn.execute("""
            SELECT table_id
            FROM user_table_permissions
            WHERE user_id = ?
        """, (user_id,)).fetchall()
    ]

    conn.close()

    return render_template(
        "user_permissions.html",
        user=user,
        tables=tables,
        allowed_table_ids=allowed_table_ids
    )


# -------------------------
# AUDIT LOGS
# -------------------------
@app.route("/audit-logs")
def audit_logs():
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    logs = conn.execute("""
        SELECT * FROM audit_logs
        ORDER BY id DESC
    """).fetchall()
    conn.close()

    return render_template("audit_logs.html", logs=logs)


if __name__ == "__main__":
=======
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
import sqlite3
from datetime import datetime, date
from werkzeug.security import check_password_hash
from openpyxl import Workbook
from io import BytesIO
from openpyxl.utils import get_column_letter
import json
import re
from datetime import datetime, date, timedelta

app = Flask(__name__)
app.secret_key = "your_secret_key_here"
DB_NAME = "company_data.db"


def get_db_connection():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def login_required():
    return "admin_id" in session


def get_expiry_status(expiry_date_str):
    try:
        expiry_date = datetime.strptime(expiry_date_str, "%Y-%m-%d").date()
        today = date.today()
        days_left = (expiry_date - today).days

        if days_left < 0:
            return "Expired", "danger", days_left
        elif days_left <= 7:
            return "Expiring Soon", "warning", days_left
        elif days_left <= 30:
            return "Near Expiry", "info", days_left
        else:
            return "Active", "success", days_left
    except Exception:
        return None, None, None


def get_alerts():
    conn = get_db_connection()

    query = """
    SELECT 
        c.company_name,
        ct.table_name,
        tr.id AS row_id,
        tc.column_name,
        rv.value AS expiry_value
    FROM row_values rv
    JOIN table_rows tr ON rv.row_id = tr.id
    JOIN table_columns tc ON rv.column_id = tc.id
    JOIN company_tables ct ON tr.table_id = ct.id
    JOIN companies c ON ct.company_id = c.id
    WHERE tc.is_expiry = 1
    """
    rows = conn.execute(query).fetchall()
    conn.close()

    alerts = []
    for row in rows:
        status, badge, days_left = get_expiry_status(row["expiry_value"])
        if status in ["Expired", "Expiring Soon", "Near Expiry"]:
            alerts.append({
                "company_name": row["company_name"],
                "table_name": row["table_name"],
                "column_name": row["column_name"],
                "expiry_value": row["expiry_value"],
                "status": status,
                "badge": badge,
                "days_left": days_left
            })

    alerts.sort(key=lambda x: x["days_left"] if x["days_left"] is not None else 99999)
    return alerts


@app.route("/")
def home():
    return render_template("home.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()

        conn = get_db_connection()
        admin = conn.execute(
            "SELECT * FROM admins WHERE username = ?",
            (username,)
        ).fetchone()
        conn.close()

        if admin and check_password_hash(admin["password_hash"], password):
            session["admin_id"] = admin["id"]
            session["username"] = admin["username"]
            flash("Login successful.", "success")
            return redirect(url_for("dashboard"))
        else:
            flash("Invalid username or password.", "danger")

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Logged out successfully.", "info")
    return redirect(url_for("login"))


@app.route("/dashboard")
def dashboard():
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    total_companies = conn.execute("SELECT COUNT(*) AS count FROM companies").fetchone()["count"]
    total_tables = conn.execute("SELECT COUNT(*) AS count FROM company_tables").fetchone()["count"]
    conn.close()

    alerts = get_alerts()

    expired_count = sum(1 for a in alerts if a["status"] == "Expired")
    soon_count = sum(1 for a in alerts if a["status"] == "Expiring Soon")
    near_count = sum(1 for a in alerts if a["status"] == "Near Expiry")

    return render_template(
        "dashboard.html",
        total_companies=total_companies,
        total_tables=total_tables,
        alerts=alerts,
        expired_count=expired_count,
        soon_count=soon_count,
        near_count=near_count
    )


@app.route("/companies", methods=["GET", "POST"])
def companies():
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        notes = request.form.get("notes", "").strip()

        if company_name:
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO companies (company_name, notes) VALUES (?, ?)",
                (company_name, notes)
            )
            conn.commit()

            new_company_id = cursor.lastrowid

            add_audit_log(
                action_type="Add Company",
                description=f"Added company '{company_name}'",
                company_id=new_company_id,
                company_name=company_name
            )

            flash("Company added successfully.", "success")
        else:
            flash("Company name is required.", "danger")

    companies_data = conn.execute("SELECT * FROM companies ORDER BY id DESC").fetchall()
    conn.close()

    return render_template("companies.html", companies=companies_data)


@app.route("/company/<int:company_id>")
def company_detail(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    tables = conn.execute(
        "SELECT * FROM company_tables WHERE company_id = ? ORDER BY id DESC",
        (company_id,)
    ).fetchall()

    # GET COMPANY DETAILS
    details = conn.execute(
        "SELECT * FROM company_details WHERE company_id = ? ORDER BY id DESC",
        (company_id,)
    ).fetchall()

    conn.close()

    if not company:
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    return render_template(
        "company_detail.html",
        company=company,
        tables=tables,
        details=details
    )


@app.route("/company/edit/<int:company_id>", methods=["GET", "POST"])
def edit_company(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    if request.method == "POST":
        company_name = request.form.get("company_name", "").strip()
        notes = request.form.get("notes", "").strip()

        if not company_name:
            flash("Company name is required.", "danger")
            return render_template("edit_company.html", company=company)

        old_name = company["company_name"]

        conn.execute(
            "UPDATE companies SET company_name = ?, notes = ? WHERE id = ?",
            (company_name, notes, company_id)
        )

        conn.commit()
        conn.close()

        # Audit log
        add_audit_log(
            action_type="Edit Company",
            description=f"Updated company '{old_name}' to '{company_name}'",
            company_id=company_id,
            company_name=company_name
        )

        flash("Company updated successfully.", "success")
        return redirect(url_for("companies"))

    conn.close()
    return render_template("edit_company.html", company=company)

def add_audit_log(action_type, description, company_id=None, company_name=None):
    conn = get_db_connection()

    dubai_time = (datetime.utcnow() + timedelta(hours=4)).strftime("%Y-%m-%d %H:%M:%S")

    conn.execute("""
        INSERT INTO audit_logs (
            admin_username,
            company_id,
            company_name,
            action_type,
            description,
            created_at
        )
        VALUES (?, ?, ?, ?, ?, ?)
    """, (
        session.get("username", "Unknown"),
        company_id,
        company_name,
        action_type,
        description,
        dubai_time
    ))

    conn.commit()
    conn.close()

@app.route("/company/delete/<int:company_id>")
def delete_company(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()
    # delete company details first
    conn.execute("DELETE FROM company_details WHERE company_id = ?", (company_id,))

    # delete all row values belonging to this company's tables
    conn.execute("""
        DELETE FROM row_values
        WHERE row_id IN (
            SELECT id FROM table_rows
            WHERE table_id IN (
                SELECT id FROM company_tables WHERE company_id = ?
            )
        )
    """, (company_id,))

    # delete rows
    conn.execute("""
        DELETE FROM table_rows
        WHERE table_id IN (
            SELECT id FROM company_tables WHERE company_id = ?
        )
    """, (company_id,))

    # delete columns
    conn.execute("""
        DELETE FROM table_columns
        WHERE table_id IN (
            SELECT id FROM company_tables WHERE company_id = ?
        )
    """, (company_id,))

    # delete tables
    conn.execute("DELETE FROM company_tables WHERE company_id = ?", (company_id,))

    # delete company
    conn.execute("DELETE FROM companies WHERE id = ?", (company_id,))
    company_name = company["company_name"] if company else "Unknown"
    conn.commit()
    conn.close()

    flash("Company deleted successfully.", "warning")
    return redirect(url_for("companies"))

    add_audit_log(
        action_type="Delete Company",
        description=f"Deleted company '{company_name}' and all related data",
        company_id=company_id,
        company_name=company_name
    )
@app.route("/company/<int:company_id>/details", methods=["GET", "POST"])
def company_details(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    if request.method == "POST":
        detail_label = request.form.get("detail_label", "").strip()
        detail_description = request.form.get("detail_description", "").strip()

        if detail_label:
            conn.execute("""
                INSERT INTO company_details (company_id, detail_label, detail_value)
                VALUES (?, ?, ?)
            """, (company_id, detail_label, detail_description))

            conn.commit()

            # AUDIT LOG
            add_audit_log(
                action_type="Add Company Detail",
                description=f"Added detail '{detail_label}' to company '{company['company_name']}'",
                company_id=company_id,
                company_name=company["company_name"]
            )

            flash("Company detail added successfully.", "success")
        else:
            flash("Detail label is required.", "danger")

    details = conn.execute("""
        SELECT * FROM company_details
        WHERE company_id = ?
        ORDER BY id DESC
    """, (company_id,)).fetchall()

    conn.close()

    return render_template(
        "company_details.html",
        company=company,
        details=details
    )



@app.route("/company/detail/edit/<int:detail_id>/<int:company_id>", methods=["GET", "POST"])
def edit_company_detail(detail_id, company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    detail = conn.execute(
        "SELECT * FROM company_details WHERE id = ?",
        (detail_id,)
    ).fetchone()

    if not detail:
        conn.close()
        flash("Detail not found.", "danger")
        return redirect(url_for("company_detail", company_id=company_id))

    if request.method == "POST":
        label = request.form.get("detail_label", "").strip()
        description = request.form.get("detail_description", "").strip()

        conn.execute(
            "UPDATE company_details SET detail_label = ?, detail_value = ? WHERE id = ?",
            (label, description, detail_id)
        )
        company = conn.execute(
    "SELECT * FROM companies WHERE id = ?",
    (company_id,)
).fetchone()
        conn.commit()
        conn.close()
        add_audit_log(
            action_type="Edit Company Detail",
            description=f"Updated detail '{label}' in company '{company['company_name']}'",
            company_id=company_id,
            company_name=company["company_name"]
        )
        flash("Company detail updated successfully.", "success")
        return redirect(url_for("company_detail", company_id=company_id))

    conn.close()
    return render_template(
        "edit_company_detail.html",
        detail=detail,
        company_id=company_id
    )

@app.route("/company/detail/delete/<int:detail_id>/<int:company_id>")
def delete_company_detail(detail_id, company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    detail = conn.execute(
        "SELECT * FROM company_details WHERE id = ?",
        (detail_id,)
    ).fetchone()

    company = conn.execute(
    "SELECT * FROM companies WHERE id = ?",
    (company_id,)
).fetchone()

    if not detail:
        conn.close()
        flash("Company detail not found.", "danger")
        return redirect(url_for("company_detail", company_id=company_id))

    conn.execute(
        "DELETE FROM company_details WHERE id = ?",
        (detail_id,)
    )
    add_audit_log(
    action_type="Delete Company Detail",
    description=f"Deleted detail '{detail['detail_label']}' from company '{company['company_name']}'",
    company_id=company_id,
    company_name=company["company_name"]
)
    conn.commit()
    conn.close()

    flash("Company detail deleted successfully.", "warning")
    return redirect(url_for("company_detail", company_id=company_id))

@app.route("/company/<int:company_id>/create_table", methods=["GET", "POST"])
def create_table(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    if request.method == "POST":
        table_name = request.form.get("table_name", "").strip()
        column_names = request.form.getlist("column_name[]")
        column_types = request.form.getlist("column_type[]")
        expiry_columns = request.form.getlist("is_expiry[]")

        if not table_name:
            flash("Table name is required.", "danger")
            conn.close()
            return render_template("create_table.html", company=company)

        cursor = conn.cursor()
        cursor.execute(
            "INSERT INTO company_tables (company_id, table_name) VALUES (?, ?)",
            (company_id, table_name)
        )
        table_id = cursor.lastrowid

        display_order = 1
        for i in range(len(column_names)):
            name = column_names[i].strip()
            col_type = column_types[i].strip() if i < len(column_types) else "text"
            is_expiry = 1 if str(i) in expiry_columns else 0

            if name:
                cursor.execute("""
                    INSERT INTO table_columns (table_id, column_name, column_type, is_expiry, display_order)
                    VALUES (?, ?, ?, ?, ?)
                """, (table_id, name, col_type, is_expiry, display_order))
                display_order += 1

        conn.commit()
        add_audit_log(
    action_type="Create Table",
    description=f"Created table '{table_name}' in company '{company['company_name']}'",
    company_id=company_id,
    company_name=company["company_name"]
)
        conn.close()

        flash("Dynamic table created successfully.", "success")
        return redirect(url_for("company_detail", company_id=company_id))

    conn.close()
    return render_template("create_table.html", company=company)

@app.route("/company/<int:company_id>/export-full")
def export_full_company(company_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    # Get company
    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (company_id,)
    ).fetchone()

    if not company:
        conn.close()
        flash("Company not found.", "danger")
        return redirect(url_for("companies"))

    workbook = Workbook()

    # ----------------------------
    # Sheet 1: Company Info
    # ----------------------------
    info_sheet = workbook.active
    info_sheet.title = "Company Info"

    info_sheet["A1"] = "Field"
    info_sheet["B1"] = "Value"

    row_num = 2
    for key in company.keys():
        info_sheet.cell(row=row_num, column=1, value=key)
        info_sheet.cell(row=row_num, column=2, value=company[key])
        row_num += 1

    # Adjust width
    info_sheet.column_dimensions["A"].width = 25
    info_sheet.column_dimensions["B"].width = 40

    # ----------------------------
    # Sheet 2: Company Details
    # ----------------------------
    details_sheet = workbook.create_sheet("Company Details")

    try:
        details = conn.execute("""
            SELECT * FROM company_details
            WHERE company_id = ?
            ORDER BY id DESC
        """, (company_id,)).fetchall()

        if details:
            headers = details[0].keys()
            for col_num, header in enumerate(headers, 1):
                details_sheet.cell(row=1, column=col_num, value=header)

            for row_index, detail in enumerate(details, start=2):
                for col_num, header in enumerate(headers, 1):
                    details_sheet.cell(row=row_index, column=col_num, value=detail[header])
        else:
            details_sheet["A1"] = "No company details found"

    except sqlite3.OperationalError:
        details_sheet["A1"] = "company_details table does not exist"

    # Adjust width for details sheet
    for col in details_sheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        details_sheet.column_dimensions[column_letter].width = max_length + 2

    # ----------------------------
    # Company dynamic tables
    # ----------------------------
    tables = conn.execute("""
        SELECT * FROM company_tables
        WHERE company_id = ?
        ORDER BY id DESC
    """, (company_id,)).fetchall()

    if not tables:
        note_sheet = workbook.create_sheet("Tables")
        note_sheet["A1"] = "No company tables found"
    else:
        for table in tables:
            table_name = table["table_name"] if table["table_name"] else f"Table_{table['id']}"
            safe_sheet_name = re.sub(r'[\\/*?:\[\]]', '', table_name)[:31]

            if not safe_sheet_name:
                safe_sheet_name = f"Table_{table['id']}"

            sheet = workbook.create_sheet(safe_sheet_name)

            # Get columns for this table
            columns = conn.execute("""
                SELECT * FROM table_columns
                WHERE table_id = ?
                ORDER BY display_order ASC
            """, (table["id"],)).fetchall()

            if not columns:
                sheet["A1"] = "No columns found for this table"
                continue

            # Header row
            headers = [column["column_name"] for column in columns]
            headers.append("Status")
            sheet.append(headers)

            # Get rows
            rows = conn.execute("""
                SELECT * FROM table_rows
                WHERE table_id = ?
                ORDER BY id DESC
            """, (table["id"],)).fetchall()

            for row in rows:
                values = conn.execute("""
                    SELECT rv.value, tc.column_name, tc.is_expiry
                    FROM row_values rv
                    JOIN table_columns tc ON rv.column_id = tc.id
                    WHERE rv.row_id = ?
                    ORDER BY tc.display_order ASC
                """, (row["id"],)).fetchall()

                row_dict = {}
                row_status = ""

                for value in values:
                    row_dict[value["column_name"]] = value["value"]
                    if value["is_expiry"] == 1:
                        status, _, _ = get_expiry_status(value["value"])
                        row_status = status if status else ""

                excel_row = [row_dict.get(column["column_name"], "") for column in columns]
                excel_row.append(row_status)
                sheet.append(excel_row)

            # Auto width
            for col in sheet.columns:
                max_length = 0
                column_letter = col[0].column_letter
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2

    conn.close()

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    company_name = company["company_name"].replace(" ", "_")
    filename = f"{company_name}_full_export.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/table/<int:table_id>")
def view_table(table_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (table["company_id"],)
    ).fetchone()

    columns = conn.execute("""
        SELECT * FROM table_columns
        WHERE table_id = ?
        ORDER BY display_order ASC
    """, (table_id,)).fetchall()

    rows = conn.execute("""
        SELECT * FROM table_rows
        WHERE table_id = ?
        ORDER BY id DESC
    """, (table_id,)).fetchall()

    table_data = []
    for row in rows:
        values = conn.execute("""
            SELECT rv.value, tc.column_name, tc.is_expiry
            FROM row_values rv
            JOIN table_columns tc ON rv.column_id = tc.id
            WHERE rv.row_id = ?
            ORDER BY tc.display_order ASC
        """, (row["id"],)).fetchall()

        row_dict = {}
        row_status = None
        row_badge = None

        for value in values:
            row_dict[value["column_name"]] = value["value"]
            if value["is_expiry"] == 1:
                status, badge, _ = get_expiry_status(value["value"])
                row_status = status
                row_badge = badge

        table_data.append({
            "id": row["id"],
            "row_values": row_dict,
            "status": row_status,
            "badge": row_badge
        })

    conn.close()

    return render_template(
        "view_table.html",
        table=table,
        company=company,
        columns=columns,
        table_data=table_data
    )



@app.route("/table/<int:table_id>/export")
def export_table_excel(table_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (table["company_id"],)
    ).fetchone()

    columns = conn.execute("""
        SELECT * FROM table_columns
        WHERE table_id = ?
        ORDER BY display_order ASC
    """, (table_id,)).fetchall()

    rows = conn.execute("""
        SELECT * FROM table_rows
        WHERE table_id = ?
        ORDER BY id DESC
    """, (table_id,)).fetchall()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Table Data"

    # Header row
    headers = [column["column_name"] for column in columns]
    headers.append("Status")
    worksheet.append(headers)

    # Data rows
    for row in rows:
        values = conn.execute("""
            SELECT rv.value, tc.column_name, tc.is_expiry
            FROM row_values rv
            JOIN table_columns tc ON rv.column_id = tc.id
            WHERE rv.row_id = ?
            ORDER BY tc.display_order ASC
        """, (row["id"],)).fetchall()

        row_dict = {}
        row_status = ""

        for value in values:
            row_dict[value["column_name"]] = value["value"]
            if value["is_expiry"] == 1:
                status, _, _ = get_expiry_status(value["value"])
                row_status = status if status else ""

        excel_row = [row_dict.get(column["column_name"], "") for column in columns]
        excel_row.append(row_status)
        worksheet.append(excel_row)

    conn.close()

    # Optional: adjust column widths a bit
    for col in worksheet.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        worksheet.column_dimensions[column_letter].width = max_length + 2

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    company_name = company["company_name"].replace(" ", "_") if company else "Company"
    table_name = table["table_name"].replace(" ", "_")
    filename = f"{company_name}_{table_name}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/table/<int:table_id>/add_row", methods=["GET", "POST"])
def add_row(table_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    columns = conn.execute("""
        SELECT * FROM table_columns
        WHERE table_id = ?
        ORDER BY display_order ASC
    """, (table_id,)).fetchall()

    if request.method == "POST":
        cursor = conn.cursor()
        cursor.execute("INSERT INTO table_rows (table_id) VALUES (?)", (table_id,))
        row_id = cursor.lastrowid

        for column in columns:
            value = request.form.get(f"column_{column['id']}", "").strip()
            cursor.execute("""
                INSERT INTO row_values (row_id, column_id, value)
                VALUES (?, ?, ?)
            """, (row_id, column["id"], value))

        conn.commit()
        company = conn.execute(
    "SELECT * FROM companies WHERE id = ?",
    (table["company_id"],)
).fetchone()
        conn.close()
        add_audit_log(
    action_type="Add Row",
    description=f"Added a row to table '{table['table_name']}' in company '{company['company_name']}'",
    company_id=company["id"],
    company_name=company["company_name"]
)
        flash("Row added successfully.", "success")
        return redirect(url_for("view_table", table_id=table_id))

    conn.close()
    return render_template("add_row.html", table=table, columns=columns)


@app.route("/row/delete/<int:row_id>/<int:table_id>")
def delete_row(row_id, table_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    table = conn.execute(
    "SELECT * FROM company_tables WHERE id = ?",
    (table_id,)
).fetchone()

    company = conn.execute(
        "SELECT * FROM companies WHERE id = ?",
        (table["company_id"],)
    ).fetchone()
    conn.execute("DELETE FROM row_values WHERE row_id = ?", (row_id,))
    conn.execute("DELETE FROM table_rows WHERE id = ?", (row_id,))
    add_audit_log(
    action_type="Delete Row",
    description=f"Deleted a row from table '{table['table_name']}' in company '{company['company_name']}'",
    company_id=company["id"],
    company_name=company["company_name"]
)
    conn.commit()
    conn.close()

    flash("Row deleted successfully.", "warning")
    return redirect(url_for("view_table", table_id=table_id))


@app.route("/table/delete/<int:table_id>")
def delete_table(table_id):
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()

    table = conn.execute(
        "SELECT * FROM company_tables WHERE id = ?",
        (table_id,)
    ).fetchone()

    if not table:
        conn.close()
        flash("Table not found.", "danger")
        return redirect(url_for("dashboard"))

    company_id = table["company_id"]

    conn.execute("""
        DELETE FROM row_values
        WHERE row_id IN (
            SELECT id FROM table_rows WHERE table_id = ?
        )
    """, (table_id,))

    conn.execute("DELETE FROM table_rows WHERE table_id = ?", (table_id,))
    conn.execute("DELETE FROM table_columns WHERE table_id = ?", (table_id,))
    conn.execute("DELETE FROM company_tables WHERE id = ?", (table_id,))

    conn.commit()
    company = conn.execute(
    "SELECT * FROM companies WHERE id = ?",
    (company_id,)
).fetchone()
    conn.close()
    add_audit_log(
    action_type="Delete Table",
    description=f"Deleted table '{table['table_name']}' from company '{company['company_name']}'",
    company_id=company_id,
    company_name=company["company_name"]
)
    flash("Table deleted successfully.", "warning")
    return redirect(url_for("company_detail", company_id=company_id))

@app.route("/audit-logs")
def audit_logs():
    if not login_required():
        return redirect(url_for("login"))

    conn = get_db_connection()
    logs = conn.execute("""
        SELECT * FROM audit_logs
        ORDER BY id DESC
    """).fetchall()
    conn.close()

    return render_template("audit_logs.html", logs=logs)

if __name__ == "__main__":
>>>>>>> 64bfbef9fa20762ecf5499a190c88bf784a8ae4f
    app.run(debug=True)